from __future__ import annotations

from pathlib import Path

from langchain_core.tools import tool


def make_list_data_files_tool(base_dir: Path):
    resolved_base = base_dir.resolve()
    data_dir = resolved_base / "data"

    @tool
    def list_data_files() -> str:
        """List all data files (Excel/CSV) in the data/ directory and return their structure:
        file name, size, sheet names, column names, and row count for each sheet.
        Call this before analyzing data to understand what files and columns are available."""
        if not data_dir.exists():
            return "list_data_files: data/ directory does not exist."

        supported = {".xlsx", ".xls", ".csv"}
        # 递归扫描所有子目录，跳过超大文件(>50MB)
        files = [f for f in data_dir.rglob("*") 
                 if f.is_file() and f.suffix.lower() in supported 
                 and f.stat().st_size < 50 * 1024 * 1024]

        if not files:
            return "list_data_files: no Excel or CSV files found in data/."

        try:
            import pandas as pd
        except ImportError:
            return "list_data_files error: pandas is not installed. Run `pip install pandas openpyxl`."

        lines: list[str] = []
        for f in sorted(files)[:30]:  # 限制最多显示 30 个文件
            size_kb = f.stat().st_size / 1024
            size_str = f"{size_kb:.1f} KB" if size_kb < 1024 else f"{size_kb / 1024:.1f} MB"
            lines.append(f"{f.name} ({size_str})")

            try:
                if f.suffix.lower() in (".xlsx", ".xls"):
                    # 使用 openpyxl 快速获取维度，不加载全部数据
                    from openpyxl import load_workbook
                    wb = load_workbook(f, read_only=True, data_only=True)
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        rows = ws.max_row or 0
                        cols = ws.max_column or 0
                        # 只显示前 10 列名
                        col_names = [str(ws.cell(1, i).value or f"Col{i}") for i in range(1, min(cols + 1, 11))]
                        lines.append(f"  [{sheet}] ~{rows} 行, 列({cols}): {', '.join(col_names)}...")
                    wb.close()
                elif f.suffix.lower() == ".csv":
                    df_head = pd.read_csv(f, nrows=0, encoding="utf-8-sig")
                    row_count = sum(1 for _ in open(f, encoding="utf-8-sig")) - 1
                    cols = ", ".join(str(c) for c in df_head.columns.tolist()[:10])
                    lines.append(f"  {row_count} 行, 列: {cols}...")
            except Exception as e:
                lines.append(f"  (读取失败: {e})")

        return "\n".join(lines)

    return list_data_files
