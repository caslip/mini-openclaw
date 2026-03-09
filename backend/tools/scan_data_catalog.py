"""
数据目录扫描脚本 (优化版)
快速扫描 data/ 目录下所有数据文件，提取基本信息
生成 data_catalog.py 和 data_catalog.md
"""

import os
import re
from pathlib import Path
from datetime import datetime
import pandas as pd

# 配置 - 使用绝对路径
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_PY = BASE_DIR / "data_catalog.py"
OUTPUT_MD = BASE_DIR / "data_catalog.md"


def infer_dtype(pandas_dtype):
    """将 pandas dtype 转换为简单类型"""
    dtype_str = str(pandas_dtype).lower()
    if "int" in dtype_str:
        return "int"
    elif "float" in dtype_str:
        return "float"
    elif "datetime" in dtype_str:
        return "datetime"
    elif "object" in dtype_str:
        return "string"
    elif "bool" in dtype_str:
        return "boolean"
    else:
        return "string"


def scan_excel_file_fast(file_path: Path) -> dict:
    """快速扫描单个 Excel 文件 - 只获取 sheet 名称和预估行数"""
    result = {
        "original_name": file_path.name,
        "path": str(file_path.relative_to(BASE_DIR)),
        "size_bytes": file_path.stat().st_size,
        "modified_time": datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
        "category": infer_category(file_path),
        "sheets": {}
    }
    
    try:
        # 使用 openpyxl 只读取 sheet 名称，不加载数据
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                # 快速获取维度（不读取所有数据）
                rows = ws.max_row or 0
                cols = ws.max_column or 0
                
                # 获取列名（第一行）
                columns = {}
                if rows > 0:
                    for col_idx in range(1, min(cols + 1, 50)):  # 最多读取50列
                        cell = ws.cell(row=1, column=col_idx)
                        col_name = str(cell.value) if cell.value else f"Column_{col_idx}"
                        columns[col_name] = {"dtype": "unknown"}
                
                result["sheets"][sheet_name] = {
                    "rows": rows,
                    "columns": columns,
                    "column_count": cols
                }
            except Exception as e:
                result["sheets"][sheet_name] = {
                    "error": str(e)
                }
        
        wb.close()
        
    except Exception as e:
        result["error"] = str(e)
    
    return result


def infer_category(file_path: Path) -> str:
    """根据路径推断数据类别"""
    path_str = str(file_path)
    if "10_Sell-in" in path_str or "Sell-in" in path_str:
        return "sell-in"
    elif "20_Sell-out" in path_str or "Sell-out" in path_str:
        return "sell-out"
    elif "30_Promo" in path_str or "Promo" in path_str:
        return "promo"
    elif "40_Listing" in path_str or "Listing" in path_str:
        return "listing"
    else:
        return "other"


def generate_catalog_key(file_path: Path) -> str:
    """根据文件路径生成 catalog key"""
    name = file_path.stem
    name = re.sub(r'\s*\(\d+\)\s*', '', name)
    name = re.sub(r'[\s\-]+', '_', name)
    name = re.sub(r'[^\w_]', '', name)
    return name


def scan_all_files():
    """扫描所有数据文件"""
    catalog = {}
    
    exts = ['.xlsx', '.xls']
    
    for ext in exts:
        for file_path in DATA_DIR.rglob(f"*{ext}"):
            if '~' in file_path.name or file_path.name.startswith('.'):
                continue
            
            key = generate_catalog_key(file_path)
            print(f"Scanning: {file_path.relative_to(BASE_DIR)}")
            
            try:
                catalog[key] = scan_excel_file_fast(file_path)
            except Exception as e:
                print(f"Error scanning {file_path.name}: {e}")
                catalog[key] = {
                    "original_name": file_path.name,
                    "path": str(file_path.relative_to(BASE_DIR)),
                    "error": str(e)
                }
    
    return catalog


def generate_python_file(catalog: dict) -> str:
    """生成 Python 格式的 catalog 文件"""
    lines = [
        '"""',
        "数据目录 Catalog",
        f"生成时间: {datetime.now().isoformat()}",
        '"""',
        "",
        "CATALOG = {",
    ]
    
    for key, info in catalog.items():
        lines.append(f'    "{key}": {{')
        lines.append(f'        "original_name": "{info.get("original_name", "")}",')
        lines.append(f'        "path": "{info.get("path", "")}",')
        lines.append(f'        "size_bytes": {info.get("size_bytes", 0)},')
        lines.append(f'        "category": "{info.get("category", "other")}",')
        
        if "sheets" in info and info["sheets"]:
            lines.append('        "sheets": {')
            for sheet_name, sheet_info in info["sheets"].items():
                if isinstance(sheet_info, dict) and "error" not in sheet_info:
                    lines.append(f'            "{sheet_name}": {{')
                    lines.append(f'                "rows": {sheet_info.get("rows", 0)},')
                    lines.append(f'                "column_count": {sheet_info.get("column_count", 0)},')
                    lines.append('                "columns": {')
                    for col_name, col_info in list(sheet_info.get("columns", {}).items())[:20]:  # 限制列数
                        lines.append(f'                    "{col_name}": "{col_info.get("dtype", "unknown")}",')
                    lines.append('                },')
                    lines.append('            },')
            lines.append('        },')
        
        lines.append('    },')
        lines.append('')
    
    lines.append('}')
    
    return '\n'.join(lines)


def generate_markdown_file(catalog: dict) -> str:
    """生成 Markdown 格式的 catalog 文件"""
    lines = [
        "# 数据目录 Catalog",
        "",
        f"生成时间: {datetime.now().isoformat()}",
        "",
        "---\n",
        "",
        "## 文件列表\n",
    ]
    
    # 按类别分组
    categories = {}
    for key, info in catalog.items():
        cat = info.get("category", "other")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append((key, info))
    
    for cat in ["sell-in", "sell-out", "promo", "listing", "other"]:
        if cat not in categories:
            continue
        
        lines.append(f"## {cat.upper()}")
        lines.append("")
        
        for key, info in categories[cat]:
            lines.append(f"### {key}")
            lines.append("")
            lines.append(f"- **原始名称**: {info.get('original_name', '')}")
            lines.append(f"- **路径**: `{info.get('path', '')}`")
            
            size_bytes = info.get('size_bytes', 0)
            if size_bytes > 1024 * 1024:
                size_str = f"{size_bytes / 1024 / 1024:.1f} MB"
            elif size_bytes > 1024:
                size_str = f"{size_bytes / 1024:.1f} KB"
            else:
                size_str = f"{size_bytes} B"
            lines.append(f"- **大小**: {size_str}")
            
            if "sheets" in info and info["sheets"]:
                for sheet_name, sheet_info in info["sheets"].items():
                    if isinstance(sheet_info, dict) and "error" not in sheet_info:
                        lines.append(f"\n#### Sheet: {sheet_name}")
                        lines.append(f"- 行数: {sheet_info.get('rows', 0):,}")
                        lines.append(f"- 列数: {sheet_info.get('column_count', 0)}")
                        
                        cols = sheet_info.get("columns", {})
                        if cols:
                            lines.append("\n**列名（前20个）:**")
                            lines.append("| 列名 | 类型 |")
                            lines.append("|------|------|")
                            for col_name, col_info in list(cols.items())[:20]:
                                lines.append(f"| {col_name} | {col_info.get('dtype', 'unknown')} |")
            
            lines.append("")
    
    return '\n'.join(lines)


if __name__ == "__main__":
    print(f"Scanning data directory: {DATA_DIR}")
    print("-" * 50)
    
    catalog = scan_all_files()
    
    print("-" * 50)
    print(f"Total files scanned: {len(catalog)}")
    
    # 生成 Python 文件
    py_content = generate_python_file(catalog)
    OUTPUT_PY.write_text(py_content, encoding='utf-8')
    print(f"Generated: {OUTPUT_PY}")
    
    # 生成 Markdown 文件
    md_content = generate_markdown_file(catalog)
    OUTPUT_MD.write_text(md_content, encoding='utf-8')
    print(f"Generated: {OUTPUT_MD}")
    
    print("Done!")
